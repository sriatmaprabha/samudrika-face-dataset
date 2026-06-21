"""
Samudrika Lakshana — Extract meaningful terms and insights from palmistry texts.
Generates one Excel sheet per book, one row per insight.
"""
import re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Key palmistry terms to watch for ──────────────────────────
TERMS = [
    "Life Line", "Line of Life", "Heart Line", "Line of Heart",
    "Head Line", "Line of Head", "Fate Line", "Line of Fate",
    "Sun Line", "Line of Sun", "Health Line", "Marriage Line",
    "Children Line", "Girdle of Venus", "Line of Mars",
    "Mount of Venus", "Mount of Jupiter", "Mount of Saturn",
    "Mount of Apollo", "Mount of Mercury", "Mount of Moon",
    "Mount of Mars",
    "Thumb", "Index finger", "Middle finger", "Ring finger",
    "Little finger", "Finger", "Nail", "Knot",
    "Square hand", "Spatulate", "Conic hand", "Psychic hand",
    "Elementary hand", "Philosophic hand",
    "Island", "Triangle", "Cross", "Star", "Grille", "Square",
    "Fork", "Trident", "Fish", "Flag", "Conch",
    "Cheirognomy", "Cheiromancy", "Hasta", "Samudrika",
    "Karma", "Destiny", "Fate", "Free will",
    "Yav", "Rascette", "Bracelet",
]

# Build a case-insensitive regex
TERM_PATTERN = re.compile(
    r'\b(' + '|'.join(re.escape(t) for t in TERMS) + r')\b',
    re.IGNORECASE
)

def clean(text):
    """Remove OCR noise and normalise whitespace."""
    text = re.sub(r'[^\x00-\x7Fऀ-ॿ -~]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    # Remove lines that are mostly noise (short, lots of numbers/symbols)
    return text

def extract_sentences(text):
    """Split text into sentences / logical units."""
    # Split on sentence-ending punctuation
    parts = re.split(r'(?<=[.!?])\s+', text)
    return [p.strip() for p in parts if len(p.strip()) > 40]

def identify_term(sentence):
    """Return the first matched palmistry term in a sentence."""
    m = TERM_PATTERN.search(sentence)
    return m.group(0).title() if m else None

def make_insight(sentence, term):
    """
    Extract a pithy insight from the sentence.
    We keep the sentence as the 'Original Text' and produce
    a shorter 'Key Insight' summary from it.
    """
    # Truncate very long sentences
    s = sentence[:300]
    # For the insight column, we'll summarise what the text says about the term
    # Simple heuristic: take the clause that contains the term
    clauses = re.split(r'[,;]', s)
    for c in clauses:
        if TERM_PATTERN.search(c):
            return c.strip()
    return s[:200]

# ─── Style helpers ─────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="800020")  # maroon
ALT_FILL      = PatternFill("solid", fgColor="FFF8E8")  # cream
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT   = Font(size=10)
WRAP          = Alignment(wrap_text=True, vertical="top")
THIN          = Side(border_style="thin", color="DDDDDD")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_sheet(ws, col_widths):
    # Header row
    for cell in ws[1]:
        cell.font        = HEADER_FONT
        cell.fill        = HEADER_FILL
        cell.alignment   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border      = BORDER
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for cell in row:
            cell.fill      = fill
            cell.font      = NORMAL_FONT
            cell.alignment = WRAP
            cell.border    = BORDER
        ws.row_dimensions[row_idx].height = 60

    # Column widths
    for col_num, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.freeze_panes = "A2"

# ─── Process one PDF ───────────────────────────────────────────
def process_book(pdf_path, sheet_name, wb, max_pages=40):
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        limit = min(max_pages, total)
        print(f"  Reading {limit} of {total} pages from: {sheet_name}")
        for page_num in range(limit):
            raw = pdf.pages[page_num].extract_text()
            if not raw:
                continue
            text = clean(raw)
            sentences = extract_sentences(text)
            for sent in sentences:
                term = identify_term(sent)
                if not term:
                    continue
                insight = make_insight(sent, term)
                rows.append({
                    "page":    page_num + 1,
                    "term":    term,
                    "text":    sent[:400],
                    "insight": insight,
                })

    if not rows:
        print(f"  ⚠ No meaningful rows found for: {sheet_name}")
        return

    ws = wb.create_sheet(title=sheet_name[:31])
    headers = ["Page", "Term / Concept", "Original Text", "Key Insight"]
    ws.append(headers)
    for r in rows:
        ws.append([r["page"], r["term"], r["text"], r["insight"]])

    style_sheet(ws, col_widths=[7, 22, 60, 45])
    print(f"  ✓ {len(rows)} rows written to sheet '{sheet_name}'")

# ─── Main ──────────────────────────────────────────────────────
def main():
    output = "/Users/sriatmaprabha/Desktop/Samudrika_Insights_SAMPLE.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    books = [
        (
            "/Users/sriatmaprabha/Desktop/samudrika_lakshana/hasta samudrika lakshana/2015.125592.Hast-Samudrika-Shastra_text.pdf",
            "Hast Samudrika Shastra",
        ),
    ]

    for path, name in books:
        print(f"\nProcessing: {name}")
        process_book(path, name, wb, max_pages=35)

    wb.save(output)
    print(f"\n✅ Sample saved to: {output}")

if __name__ == "__main__":
    main()
