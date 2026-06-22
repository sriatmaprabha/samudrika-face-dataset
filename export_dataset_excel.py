"""
export_dataset_excel.py
Converts dataset.jsonl → Excel for review.
One row per image, one column per feature.
"""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).parent
JSONL = HERE / "dataset.jsonl"
OUT   = HERE / "data_sets/output/Samudrika_Face_Dataset.xlsx"

records = [json.loads(l) for l in JSONL.read_text().splitlines() if l.strip()]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Face Readings"

HEADER_FILL  = PatternFill("solid", fgColor="1a1a4a")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
MALE_FILL    = PatternFill("solid", fgColor="dce8f5")
FEMALE_FILL  = PatternFill("solid", fgColor="f5dce8")
AUSP_FILL    = PatternFill("solid", fgColor="d4edda")
INAUSP_FILL  = PatternFill("solid", fgColor="f8d7da")
MIXED_FILL   = PatternFill("solid", fgColor="fff3cd")

FEATURES = ['face_shape', 'forehead', 'eyebrows', 'eyes', 'nose',
            'lips_mouth', 'chin', 'ears', 'complexion']

# Build headers
headers = [
    "File", "Gender", "Examination Side",
]
for feat in FEATURES:
    label = feat.replace('_', ' ').title()
    headers += [f"{label} — Type/Shape", f"{label} — Observation",
                f"{label} — Shastra Meaning", f"{label} — Quality"]

headers += ["Overall Reading", "Prediction 1", "Prediction 2", "Prediction 3",
            "Dominant Quality", "Confidence Note"]

ws.append(headers)

# Style header row
for col, cell in enumerate(ws[1], 1):
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[1].height = 40

# Fill data
for rec in records:
    a = rec["analysis"]
    row = [
        rec["file"],
        rec["gender"].capitalize(),
        a.get("examination_side", ""),
    ]
    for feat in FEATURES:
        d = a.get(feat, {})
        # type field varies by feature name
        feat_type = (d.get("type") or d.get("shape") or d.get("tone") or
                     d.get("lines_count") or d.get("size") or "")
        row += [
            str(feat_type),
            d.get("observation", ""),
            d.get("shastra_meaning", ""),
            d.get("quality", ""),
        ]
    preds = a.get("key_predictions", ["", "", ""])
    while len(preds) < 3: preds.append("")
    row += [
        a.get("overall_reading", ""),
        preds[0], preds[1], preds[2],
        a.get("dominant_quality", ""),
        a.get("confidence_note", ""),
    ]
    ws.append(row)

# Style data rows
for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
    gender = ws.cell(row_idx, 2).value or ""
    row_fill = MALE_FILL if gender.lower() == "male" else FEMALE_FILL
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = row_fill

    # Colour quality cells
    for col_offset, feat in enumerate(FEATURES):
        qual_col = 4 + col_offset * 4  # quality is 4th sub-column per feature
        qual_cell = ws.cell(row_idx, qual_col)
        q = str(qual_cell.value or "").lower()
        if q == "auspicious":    qual_cell.fill = AUSP_FILL
        elif q == "inauspicious": qual_cell.fill = INAUSP_FILL
        elif q == "mixed":        qual_cell.fill = MIXED_FILL

    # Colour dominant quality
    dom_col = 4 + len(FEATURES) * 4 + 4  # after all features + 4 text cols
    dom_cell = ws.cell(row_idx, dom_col)
    q = str(dom_cell.value or "").lower()
    if q == "auspicious":    dom_cell.fill = AUSP_FILL
    elif q == "inauspicious": dom_cell.fill = INAUSP_FILL
    elif q == "mixed":        dom_cell.fill = MIXED_FILL

# Column widths
ws.column_dimensions["A"].width = 18  # File
ws.column_dimensions["B"].width = 10  # Gender
ws.column_dimensions["C"].width = 18  # Side
for col_idx in range(4, len(headers) + 1):
    letter = openpyxl.utils.get_column_letter(col_idx)
    h = headers[col_idx - 1]
    if "Observation" in h or "Meaning" in h or "Reading" in h or "Prediction" in h:
        ws.column_dimensions[letter].width = 45
    elif "Quality" in h or "Type" in h or "Shape" in h:
        ws.column_dimensions[letter].width = 16
    else:
        ws.column_dimensions[letter].width = 20

# Freeze header + first 2 cols
ws.freeze_panes = "D2"

wb.save(OUT)
print(f"✅ Saved: {OUT}")
print(f"   {len(records)} rows | {len(headers)} columns")
print(f"   Blue rows = male | Pink rows = female")
print(f"   Green = auspicious | Red = inauspicious | Yellow = mixed")
