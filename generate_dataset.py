"""
Samudrika Face Dataset Generator
Analyzes 50 face images using Nvidia vision API.
Context: sl notes (face sections) + face-relevant Excel rows.
Output: dataset.jsonl — one JSON record per image.
"""
import base64, json, re, subprocess, os, glob
from pathlib import Path
import openpyxl

API_KEY  = "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"
MODEL    = "meta/llama-3.2-90b-vision-instruct"
BASE_URL = "https://integrate.api.nvidia.com/v1"

HERE      = Path(__file__).parent
FACES_DIR = HERE / "faces"
EXCEL     = HERE / "data_sets/output/Samudrika_Insights.xlsx"
SL_NOTES  = HERE / "sl notes"
OUT       = HERE / "dataset.jsonl"
TEMP_DIR  = Path("/tmp/sl_jpegs")
TEMP_DIR.mkdir(exist_ok=True)

# ── 1. Extract face sections from sl notes ─────────────────────────────────

FACE_SECTIONS = [
    "Head", "fore head", "Eyebrows", "Eyelids", "Eyes", "Face",
    "Facial hair", "Cheek", "Nose", "Mouth", "lips", "Tooth", "Ear", "Chin"
]

def extract_sl_notes_face():
    text = SL_NOTES.read_text()
    lines = text.split('\n')
    capturing = False
    sections  = []
    buf = []
    current = ""
    STOP_AT = re.compile(r'^### (Neck|Nape|Shoulder|Arm|Wrist|Palm|Finger|Chest|Belly|Navel|Thigh|Knee|Leg|Feet|Foot|Ankle|Mole|Birth mark|Whor|Ten Dasha|measurement)', re.I)
    START_AT = re.compile(r'^### (' + '|'.join(re.escape(s) for s in FACE_SECTIONS) + r')', re.I)

    for line in lines:
        if START_AT.match(line):
            if buf and current:
                sections.append(f"**{current}**\n" + '\n'.join(buf).strip())
            current = line.strip('# ').strip()
            buf = []
            capturing = True
        elif STOP_AT.match(line) and capturing:
            break
        elif capturing:
            buf.append(line)

    if buf and current:
        sections.append(f"**{current}**\n" + '\n'.join(buf).strip())

    return '\n\n'.join(sections)

# ── 2. Extract face rows from Excel ────────────────────────────────────────

FACE_RE = re.compile(
    r'\b(forehead|eyebrow|brow|eyelid|eye[^b]|eyes|nose|nostril|'
    r'lip|mouth|chin|ear|complexion|cheek|jaw|face shape|'
    r'lalat|bhroo|nayana|nasa|oshtha|hanu|karna|mukha|anana|'
    r'ललाट|भ्रू|नेत्र|नाक|ओंठ|ठोड़|कान|मुख|चेहर|नासि)\b',
    re.IGNORECASE
)

def extract_excel_face_rows():
    wb   = openpyxl.load_workbook(str(EXCEL))
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        orig_i = next((i for i,h in enumerate(headers) if h and 'original' in str(h).lower()), None)
        exp_i  = next((i for i,h in enumerate(headers) if h and ('explanation' in str(h).lower() or 'simple' in str(h).lower())), None)
        if orig_i is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            orig = str(row[orig_i] or '').strip()
            exp  = str(row[exp_i]  or '').strip() if exp_i is not None else ''
            # Skip palm-color / hand-line rows that mention 'eye disease' only
            if not FACE_RE.search(orig) and not FACE_RE.search(exp):
                continue
            if re.search(r'\b(heart line|life line|fate line|sun line|mount of|palm line|hatheli|हथेली)\b', orig+exp, re.I):
                continue
            entry = exp if exp else orig
            if len(entry) > 20:
                rows.append(entry[:250])
    # Deduplicate
    seen, unique = set(), []
    for r in rows:
        key = r[:80]
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique

# ── 3. Build system prompt ─────────────────────────────────────────────────

def build_system_prompt():
    sl_face = extract_sl_notes_face()
    excel_rows = extract_excel_face_rows()
    excel_sample = '\n'.join(f'- {r}' for r in excel_rows[:80])  # top 80 rows

    return f"""You are an expert in Samudrika Shastra — the ancient Indian science of physiognomy (face and body reading).

CLASSICAL TEXT KNOWLEDGE — FACE FEATURES:
{sl_face}

ADDITIONAL CLASSICAL INSIGHTS (from extracted texts):
{excel_sample}

EXAMINATION RULES:
- For WOMEN: examine the LEFT side of the face
- For MEN: examine the RIGHT side of the face
- FIRST examine lifespan indicators, THEN other features
- People have MIXED features — give balanced readings

OUTPUT: Return ONLY a valid JSON object. No extra text."""

# ── 4. Image helpers ────────────────────────────────────────────────────────

def to_jpeg(path):
    p = Path(path)
    if p.suffix.lower() in ('.jpg', '.jpeg'):
        return str(p)
    out = TEMP_DIR / (p.stem + '.jpg')
    if not out.exists():
        subprocess.run(['sips', '-s', 'format', 'jpeg',
                        '-Z', '1024', str(p), '--out', str(out)],
                       capture_output=True)
    return str(out)

def to_b64(path):
    with open(path, 'rb') as f:
        return base64.b64encode(f.read()).decode()

# ── 5. Analysis schema ──────────────────────────────────────────────────────

SCHEMA = {
    "gender": "male or female",
    "examination_side": "right (male) or left (female)",
    "face_shape": {
        "type": "oval|round|long|square|broad",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "forehead": {
        "lines_count": "0-5 (count visible lines)",
        "width": "broad|medium|narrow",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "eyebrows": {
        "shape": "arched|straight|joined|thin|thick",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "eyes": {
        "shape": "broad|lotus|small|round|narrow",
        "pupil": "dark|medium|light",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "nose": {
        "type": "elevated|flat|parrot-beak|broad|sharp",
        "nostrils": "small|medium|large",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "lips_mouth": {
        "size": "small|medium|large",
        "color": "red|pink|dark|medium",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "chin": {
        "shape": "single|divided|pointed|round|square",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "ears": {
        "size": "long|medium|small",
        "lobe": "hanging|attached|medium",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "complexion": {
        "tone": "fair|wheat|medium|dark",
        "observation": "what you see",
        "shastra_meaning": "classical interpretation",
        "quality": "auspicious|inauspicious|mixed"
    },
    "overall_reading": "3-4 sentence Samudrika Shastra life reading",
    "key_predictions": ["prediction 1", "prediction 2", "prediction 3"],
    "dominant_quality": "auspicious|inauspicious|mixed",
    "confidence_note": "any caveats (beard obscures chin, glasses, etc.)"
}

# ── 6. API call ─────────────────────────────────────────────────────────────

import urllib.request

def analyze(img_path, gender, system_prompt):
    jpeg  = to_jpeg(img_path)
    b64   = to_b64(jpeg)
    side  = "LEFT" if gender == "female" else "RIGHT"

    user_msg = f"""Analyze this {gender}'s face according to Samudrika Shastra.
Examine the {side} side of the face (classical rule for {gender}s).

Return ONLY this JSON structure filled with specific observations:
{json.dumps(SCHEMA, indent=2)}"""

    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                {"type": "text", "text": user_msg}
            ]}
        ],
        "max_tokens": 2500,
        "temperature": 0.1,
    }
    req = urllib.request.Request(
        f"{BASE_URL}/chat/completions",
        data=json.dumps(payload).encode(),
        headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=120) as r:
        resp = json.loads(r.read())
    content = resp["choices"][0]["message"]["content"].strip()
    start = content.find('{')
    end   = content.rfind('}') + 1
    if start == -1 or end == 0:
        return None
    return json.loads(content[start:end])

# ── 7. Main loop ────────────────────────────────────────────────────────────

# Resume: load already-done files
done = set()
if OUT.exists():
    for line in OUT.read_text().splitlines():
        try:
            done.add(json.loads(line)["file"])
        except:
            pass

print(f"Building system prompt from sl notes + Excel...")
system_prompt = build_system_prompt()
print(f"  sl notes face sections: {len(system_prompt.split(chr(10)))} lines")

excel_rows = extract_excel_face_rows()
print(f"  Excel face rows loaded: {len(excel_rows)}")
print(f"  System prompt size: {len(system_prompt):,} chars")
print(f"Already done: {len(done)} images\n")

total_done = len(done)
total_images = 0

for gender in ["male", "female"]:
    imgs = sorted(glob.glob(str(FACES_DIR / gender / "*.*")))
    total_images += len(imgs)
    print(f"=== {gender.upper()} ({len(imgs)} images) ===")

    for img_path in imgs:
        fname = os.path.basename(img_path)
        if fname in done:
            print(f"  skip: {fname}")
            continue

        print(f"  {fname}...", end=" ", flush=True)
        try:
            analysis = analyze(img_path, gender, system_prompt)
            if not analysis:
                print("✗ no JSON in response")
                continue

            record = {
                "file":         fname,
                "gender":       gender,
                "image_path":   img_path,
                "analysis":     analysis,
                # Fine-tuning target: structured JSON output
                "target_text":  json.dumps(analysis, ensure_ascii=False),
                # Conversation format for Unsloth/Gemma4
                "conversation": [
                    {
                        "role": "user",
                        "content": f"Analyze this {gender}'s face according to Samudrika Shastra. Examine the {'left' if gender=='female' else 'right'} side."
                    },
                    {
                        "role": "assistant",
                        "content": json.dumps(analysis, ensure_ascii=False)
                    }
                ]
            }
            with open(OUT, 'a') as f:
                f.write(json.dumps(record, ensure_ascii=False) + '\n')
            done.add(fname)
            total_done += 1
            reading = analysis.get("overall_reading", "")
            print(f"✓  {reading[:80]}")

        except Exception as e:
            print(f"✗ {e}")

print(f"\n✅ Dataset complete: {total_done}/{total_images} images → {OUT}")
print(f"Next: run format_hf_dataset.py to convert to HuggingFace Parquet format")
