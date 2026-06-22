"""
test_reading.py — Test Samudrika face reading on any image.
Usage: python3 test_reading.py <image_path> <male|female>
Example: python3 test_reading.py ~/Downloads/face.jpg male
"""
import sys, json, base64, subprocess, urllib.request
from pathlib import Path
import openpyxl, re

API_KEY  = "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"
MODEL    = "meta/llama-3.2-90b-vision-instruct"
BASE_URL = "https://integrate.api.nvidia.com/v1"
HERE     = Path(__file__).parent

# ── Load sl notes face sections ─────────────────────────────────────────────
def load_sl_notes():
    text  = (HERE / "sl notes").read_text()
    lines = text.split('\n')
    FACE  = ['Head','fore head','Eyebrows','Eyelids','Eyes','Face',
             'Facial hair','Cheek','Nose','Mouth','lips','Tooth','Ear','Chin']
    STOP  = re.compile(r'^### (Neck|Nape|Shoulder|Arm|Wrist|Palm|Finger|Chest)', re.I)
    START = re.compile(r'^### (' + '|'.join(re.escape(s) for s in FACE) + r')', re.I)
    capturing, sections, buf, current = False, [], [], ''
    for line in lines:
        if START.match(line):
            if buf and current: sections.append(f"**{current}**\n" + '\n'.join(buf).strip())
            current = line.strip('# ').strip(); buf = []; capturing = True
        elif STOP.match(line) and capturing: break
        elif capturing: buf.append(line)
    if buf and current: sections.append(f"**{current}**\n" + '\n'.join(buf).strip())
    return '\n\n'.join(sections)

# ── Load Excel face rows ─────────────────────────────────────────────────────
def load_excel_rows():
    import warnings; warnings.filterwarnings('ignore')
    wb = openpyxl.load_workbook(str(HERE / 'data_sets/output/Samudrika_Insights.xlsx'))
    FACE_RE = re.compile(r'\b(forehead|eyebrow|brow|eyelid|eyes|nose|nostril|lip|mouth|chin|ear|complexion|cheek|face shape)\b', re.I)
    rows = []
    for ws in [wb[s] for s in wb.sheetnames]:
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        oi = next((i for i,h in enumerate(headers) if h and 'original' in str(h).lower()), None)
        ei = next((i for i,h in enumerate(headers) if h and ('explanation' in str(h).lower() or 'simple' in str(h).lower())), None)
        if oi is None: continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            orig = str(row[oi] or '').strip()
            exp  = str(row[ei] or '').strip() if ei else ''
            if not FACE_RE.search(orig) and not FACE_RE.search(exp): continue
            if re.search(r'\b(heart line|life line|fate line|mount of|hatheli|हथेली)\b', orig+exp, re.I): continue
            entry = exp if exp else orig
            if len(entry) > 20: rows.append(entry[:250])
    seen, unique = set(), []
    for r in rows:
        if r[:80] not in seen: seen.add(r[:80]); unique.append(r)
    return unique

# ── Convert image to JPEG base64 ─────────────────────────────────────────────
def image_to_b64(path):
    p = Path(path).expanduser()
    if p.suffix.lower() in ('.avif', '.webp'):
        tmp = Path(f"/tmp/{p.stem}_test.jpg")
        subprocess.run(['sips', '-s', 'format', 'jpeg', '-Z', '1024',
                        str(p), '--out', str(tmp)], capture_output=True)
        p = tmp
    with open(p, 'rb') as f:
        return base64.b64encode(f.read()).decode()

# ── Main ─────────────────────────────────────────────────────────────────────
def run(img_path, gender):
    print("Loading Samudrika knowledge...")
    sl   = load_sl_notes()
    rows = load_excel_rows()
    excel_context = '\n'.join(f'- {r}' for r in rows[:80])

    system = f"""You are an expert in Samudrika Shastra — the ancient Indian science of physiognomy.

CLASSICAL TEXT — FACE FEATURES:
{sl}

ADDITIONAL INSIGHTS FROM CLASSICAL TEXTS:
{excel_context}

RULES:
- For WOMEN: examine the LEFT side of the face
- For MEN: examine the RIGHT side of the face
- First assess lifespan indicators, then other features
- People have MIXED features — give balanced readings
- Return ONLY valid JSON, no other text."""

    schema = {
        "gender": gender,
        "examination_side": "right" if gender == "male" else "left",
        "face_shape":  {"type": "oval|round|long|square|broad", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "forehead":    {"lines_count": "0-5", "width": "broad|medium|narrow", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "eyebrows":    {"shape": "arched|straight|joined|thin|thick", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "eyes":        {"shape": "broad|lotus|small|round|narrow", "pupil": "dark|medium|light", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "nose":        {"type": "elevated|flat|parrot-beak|broad|sharp", "nostrils": "small|medium|large", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "lips_mouth":  {"size": "small|medium|large", "color": "red|pink|dark|medium", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "chin":        {"shape": "single|divided|pointed|round|square", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "ears":        {"size": "long|medium|small", "lobe": "hanging|attached|medium", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "complexion":  {"tone": "fair|wheat|medium|dark", "observation": "", "shastra_meaning": "", "quality": "auspicious|inauspicious|mixed"},
        "overall_reading":  "3-4 sentence Samudrika life reading",
        "key_predictions":  ["prediction 1", "prediction 2", "prediction 3"],
        "dominant_quality": "auspicious|inauspicious|mixed",
        "confidence_note":  "any caveats (beard, glasses, angle, etc.)"
    }

    prompt = f"""Analyze this {gender}'s face according to Samudrika Shastra.
Examine the {'LEFT' if gender == 'female' else 'RIGHT'} side of the face.

Return ONLY this JSON with all fields filled:
{json.dumps(schema, indent=2)}"""

    print(f"Sending to {MODEL}...")
    b64 = image_to_b64(img_path)
    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                {"type": "text", "text": prompt}
            ]}
        ],
        "max_tokens": 2500,
        "temperature": 0.1,
    }
    req = urllib.request.Request(
        f"{BASE_URL}/chat/completions",
        data=json.dumps(payload).encode(),
        headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=120) as r:
        resp = json.loads(r.read())
    content = resp["choices"][0]["message"]["content"].strip()
    start, end = content.find('{'), content.rfind('}') + 1
    result = json.loads(content[start:end])

    # Pretty print
    print("\n" + "="*60)
    print(f"SAMUDRIKA READING — {gender.upper()} — {Path(img_path).name}")
    print("="*60)
    features = ['face_shape','forehead','eyebrows','eyes','nose','lips_mouth','chin','ears','complexion']
    for feat in features:
        d = result.get(feat, {})
        obs  = d.get('observation','')[:80]
        mean = d.get('shastra_meaning','')[:80]
        qual = d.get('quality','')
        icon = '✓' if qual == 'auspicious' else ('✗' if qual == 'inauspicious' else '~')
        print(f"\n{feat.upper().replace('_',' ')} {icon}")
        print(f"  Observed : {obs}")
        print(f"  Meaning  : {mean}")

    print(f"\n{'─'*60}")
    print(f"OVERALL READING:")
    print(f"  {result.get('overall_reading','')}")
    print(f"\nKEY PREDICTIONS:")
    for p in result.get('key_predictions', []):
        print(f"  • {p}")
    print(f"\nDOMINANT QUALITY: {result.get('dominant_quality','').upper()}")
    if result.get('confidence_note'):
        print(f"NOTE: {result.get('confidence_note','')}")
    print("="*60)

    # Save full JSON
    out = Path(img_path).stem + "_reading.json"
    Path(out).write_text(json.dumps(result, indent=2, ensure_ascii=False))
    print(f"\nFull JSON saved: {out}")
    return result

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 test_reading.py <image_path> <male|female>")
        sys.exit(1)
    img_path = sys.argv[1]
    gender   = sys.argv[2].lower()
    if gender not in ('male', 'female'):
        print("Gender must be 'male' or 'female'"); sys.exit(1)
    run(img_path, gender)
