"""
build_text_dataset.py
Converts all 6,394 Excel rows into instruction-following training pairs.

3 Q&A templates per row → ~19,000 text pairs.
Output: dataset_text.jsonl  +  data_sets/output/Samudrika_Text_Dataset.xlsx
"""
import json, re, warnings
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings('ignore')

HERE  = Path(__file__).parent
EXCEL = HERE / "data_sets/output/Samudrika_Insights.xlsx"
OUT_J = HERE / "dataset_text.jsonl"
OUT_X = HERE / "data_sets/output/Samudrika_Text_Dataset.xlsx"

# Book → topic tag mapping
BOOK_TOPIC = {
    "Hast Rekha Shastra":               "palmistry",
    "Hast Rekha Shastra (v2)":          "palmistry",
    "Hast Samudrika Shastra (Indian":   "palmistry_body",
    "Hast Samudrika Shastra (Indian 1": "palmistry_body",
    "Hast Samudrika Shastra (Indian 2": "palmistry_body",
    "Hast Samudrika Shastra (v2)":      "palmistry_body",
    "Samudrak Shastra Part 1 and 2":    "body_reading",
    "Samudrika Laksanam":               "face_body",
    "Samudrika Shastra":                "face_body",
    "Saral Samudrik Shastra — Face R":  "face_reading",
}

# Detect if a term/row is face-related
FACE_RE = re.compile(
    r'\b(forehead|eyebrow|eye|nose|nostril|lip|mouth|chin|ear|complexion|'
    r'cheek|jaw|face|head|hair|neck|tooth|teeth|tongue|brow|'
    r'ललाट|भ्रू|नेत्र|नाक|ओंठ|ठोड़|कान|मुख|चेहर)\b', re.I)

PALM_RE = re.compile(
    r'\b(mount|line|palm|finger|thumb|nail|wrist|hand|jupiter|saturn|'
    r'apollo|mercury|venus|mars|luna|moon|heart line|life line|fate)\b', re.I)

def detect_topic(term, orig, exp, sheet_topic):
    combined = f"{term} {orig} {exp}"
    if FACE_RE.search(combined): return "face_body"
    if PALM_RE.search(combined): return "palmistry"
    return sheet_topic

def make_pairs(term, original, explanation, what_it_means, book, topic, page):
    """Generate 3 instruction-following Q&A pairs from one row."""
    pairs = []
    term    = (term or "").strip()
    orig    = (original or "").strip()
    exp     = (explanation or "").strip()
    meaning = (what_it_means or "").strip()

    # Skip rows with no useful content
    if not exp and not meaning: return []
    if len(exp) < 15 and len(meaning) < 15: return []

    # Combined answer = explanation + deeper meaning
    if meaning and meaning != exp:
        full_answer = f"{exp}\n\n{meaning}"
    else:
        full_answer = exp or meaning

    # ── Template 1: Direct question about the term ─────────────────────────
    if term and len(term) > 3 and not term.startswith('['):
        q1 = f"According to Samudrika Shastra, what does {term} indicate about a person?"
        pairs.append({"q": q1, "a": full_answer, "template": "term_question"})

    # ── Template 2: Explain the classical text ─────────────────────────────
    if orig and len(orig) > 20:
        # Is it Hindi/Sanskrit?
        has_deva = sum(1 for c in orig if 'ऀ' <= c <= 'ॿ') > 5
        if has_deva:
            q2 = f"Explain this Samudrika Shastra verse in simple terms:\n{orig}"
        else:
            q2 = f"What is the Samudrika Shastra interpretation of: \"{orig[:200]}\""
        pairs.append({"q": q2, "a": full_answer, "template": "verse_explanation"})

    # ── Template 3: Practical / predictive question ────────────────────────
    if exp and len(exp) > 30:
        if topic in ("face_reading", "face_body") and term:
            q3 = f"In Samudrika Shastra face reading, what are the classical signs and predictions for {term}?"
        elif topic == "palmistry" and term:
            q3 = f"In Samudrika Shastra hand reading, describe the significance of {term}."
        else:
            q3 = f"What does Samudrika Shastra teach about {term}?" if term else None
        if q3:
            pairs.append({"q": q3, "a": full_answer, "template": "practical"})

    # ── Template 4: Prediction / life outcome ──────────────────────────────
    if term and len(term) > 3 and not term.startswith('['):
        q4 = f"If a person has {term} according to Samudrika Shastra, what does it predict about their life, wealth, and character?"
        pairs.append({"q": q4, "a": full_answer, "template": "prediction"})

    # ── Template 5: Auspicious / inauspicious distinction ──────────────────
    if exp and len(exp) > 30:
        if topic in ("face_reading", "face_body"):
            q5 = f"Is {term} considered auspicious or inauspicious in Samudrika Shastra, and why?" if term else None
        elif topic in ("palmistry", "palmistry_body"):
            q5 = f"What is the auspicious or inauspicious significance of {term} in Samudrika Shastra palm reading?" if term else None
        else:
            q5 = f"According to Samudrika Shastra, what is the spiritual and practical significance of {term}?" if term else None
        if q5:
            pairs.append({"q": q5, "a": full_answer, "template": "auspicious_check"})

    # Tag every pair with metadata
    for p in pairs:
        p.update({
            "book":     book,
            "topic":    topic,
            "page":     str(page or ""),
            "term":     term,
            # Conversation format for Unsloth (text-only, no image)
            "conversation": [
                {"role": "user",      "content": p["q"]},
                {"role": "assistant", "content": p["a"]},
            ],
            # Simple text format matching LaTeX OCR style
            "text": p["a"],
            "instruction": p["q"],
        })
    return pairs

# ── Process all sheets ───────────────────────────────────────────────────────
wb = openpyxl.load_workbook(str(EXCEL))
all_pairs = []
stats = {}

for sheet_name in wb.sheetnames:
    ws      = wb[sheet_name]
    topic   = next((v for k,v in BOOK_TOPIC.items() if k in sheet_name), "samudrika")
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Map column indices
    idx = {h.lower(): i for i, h in enumerate(headers) if h}
    page_i  = next((i for h,i in idx.items() if 'page' in h), 0)
    term_i  = next((i for h,i in idx.items() if 'term' in h or 'concept' in h), 1)
    orig_i  = next((i for h,i in idx.items() if 'original' in h), 2)
    exp_i   = next((i for h,i in idx.items() if 'simple' in h or 'explanation' in h), 3)
    mean_i  = next((i for h,i in idx.items() if 'means' in h or 'what' in h), 4)

    sheet_pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(i): return str(row[i] or '').strip() if i < len(row) else ''
        page    = g(page_i)
        term    = g(term_i)
        orig    = g(orig_i)
        exp     = g(exp_i)
        meaning = g(mean_i)
        row_topic = detect_topic(term, orig, exp, topic)
        pairs = make_pairs(term, orig, exp, meaning, sheet_name, row_topic, page)
        sheet_pairs.extend(pairs)

    all_pairs.extend(sheet_pairs)
    stats[sheet_name] = len(sheet_pairs)
    print(f"  {sheet_name[:45]:47} → {len(sheet_pairs):5} pairs")

# ── Write JSONL ──────────────────────────────────────────────────────────────
with open(OUT_J, 'w') as f:
    for p in all_pairs:
        f.write(json.dumps(p, ensure_ascii=False) + '\n')

# ── Write Excel for review ───────────────────────────────────────────────────
ewb = openpyxl.Workbook()
ews = ewb.active
ews.title = "Text Training Pairs"

HFILL = PatternFill("solid", fgColor="1a1a4a")
HFONT = Font(bold=True, color="FFFFFF", size=10)
FILLS = {
    "face_reading":   PatternFill("solid", fgColor="d4edda"),
    "face_body":      PatternFill("solid", fgColor="cce5ff"),
    "body_reading":   PatternFill("solid", fgColor="fff3cd"),
    "palmistry":      PatternFill("solid", fgColor="f8d7da"),
    "palmistry_body": PatternFill("solid", fgColor="fdebd0"),
    "samudrika":      PatternFill("solid", fgColor="f0f0f0"),
}

headers_x = ["#", "Book", "Topic", "Term", "Template", "Question", "Answer"]
ews.append(headers_x)
for c in ews[1]:
    c.font = HFONT; c.fill = HFILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
ews.row_dimensions[1].height = 30

for i, p in enumerate(all_pairs, 1):
    ews.append([i, p["book"], p["topic"], p["term"], p["template"], p["q"], p["a"]])
    fill = FILLS.get(p["topic"], FILLS["samudrika"])
    for c in ews[i+1]:
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")

ews.column_dimensions["A"].width = 6
ews.column_dimensions["B"].width = 30
ews.column_dimensions["C"].width = 16
ews.column_dimensions["D"].width = 25
ews.column_dimensions["E"].width = 18
ews.column_dimensions["F"].width = 50
ews.column_dimensions["G"].width = 60
ews.freeze_panes = "F2"

ewb.save(OUT_X)

# ── Summary ──────────────────────────────────────────────────────────────────
by_topic = {}
for p in all_pairs:
    by_topic[p["topic"]] = by_topic.get(p["topic"], 0) + 1

print(f"\n{'='*55}")
print(f"TOTAL PAIRS: {len(all_pairs):,}")
print(f"{'='*55}")
print(f"\nBy topic:")
for t, n in sorted(by_topic.items(), key=lambda x: -x[1]):
    print(f"  {t:20} : {n:6,}")
print(f"\nBy template:")
for tmpl in ["term_question", "verse_explanation", "practical"]:
    n = sum(1 for p in all_pairs if p["template"] == tmpl)
    print(f"  {tmpl:22} : {n:6,}")
print(f"\nSaved:")
print(f"  {OUT_J}  ({len(all_pairs):,} pairs)")
print(f"  {OUT_X}")
