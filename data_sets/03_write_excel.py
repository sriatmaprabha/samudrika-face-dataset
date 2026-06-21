"""
Stage 3 → Excel
Reads all JSON files in data_sets/stage3/ and writes one sheet per book.
Columns: Page | Term | Original Text | Simple Explanation | What It Means
"""
import json, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE   = Path(__file__).parent
S3_DIR = BASE / "stage3"
OUT    = BASE / "output" / "Samudrika_Insights.xlsx"
OUT.parent.mkdir(exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="800020")
ALT_FILL    = PatternFill("solid", fgColor="FFF8E8")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
BODY_FONT   = Font(size=10, name="Calibri")
WRAP        = Alignment(wrap_text=True, vertical="top")
THIN        = Side(border_style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COL_WIDTHS  = [6, 26, 62, 46, 46]
HEADERS     = ["Page", "Term / Concept", "Original Text", "Simple Explanation", "What It Means"]

def style_ws(ws):
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[1].height = 30
    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for cell in row:
            cell.fill      = fill
            cell.font      = BODY_FONT
            cell.alignment = WRAP
            cell.border    = BORDER
        ws.row_dimensions[ri].height = 65
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

def main():
    files = sorted(S3_DIR.glob("*.json"))
    if not files:
        print("No stage3 files found. Run the workflow first.")
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    total_rows = 0

    for f in files:
        data = json.loads(f.read_text())
        book_name = data.get("book", f.stem)[:31]
        rows      = data.get("approved", [])
        if not rows:
            print(f"  skip (0 rows): {book_name}")
            continue

        ws = wb.create_sheet(title=book_name)
        ws.append(HEADERS)
        for r in rows:
            ws.append([
                r.get("page", ""),
                r.get("term", ""),
                r.get("original", "")[:500],
                r.get("simple_explanation", "")[:400],
                r.get("what_it_means", "")[:400],
            ])
        style_ws(ws)
        total_rows += len(rows)
        print(f"  ✓ {len(rows):>4} rows  →  {book_name}")

    wb.save(str(OUT))
    print(f"\n✅  {total_rows} total rows  →  {OUT}")

if __name__ == "__main__":
    main()
